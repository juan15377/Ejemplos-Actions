import openpyxl
from datetime import datetime

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reporte"

# Encabezados
ws['A1'] = "Mensaje"
ws['B1'] = "Fecha de generación"

# Poner en negrita los encabezados
ws['A1'].font = openpyxl.styles.Font(bold=True)
ws['B1'].font = openpyxl.styles.Font(bold=True)

# Datos
ws['A2'] = "¡Hola mundo desde GitHub Actions!"
ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Ajustar ancho de columnas para que se vea mejor
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20

# Guardar el archivo
nombre_archivo = "resultado.xlsx"
wb.save(nombre_archivo)

print(f"Archivo '{nombre_archivo}' creado exitosamente.")